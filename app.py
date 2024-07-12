from flask import Flask, request, redirect, url_for, render_template,send_file
from werkzeug.utils import secure_filename
import os
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import openpyxl

app = Flask(__name__)
UPLOAD_FOLDER = '\OCRfile'  # 修改为你希望保存文件的路径
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Load environment variables

# Retrieve and validate environment variables
AZURE_API_ENDPOINT = os.getenv("AZURE_FORM_RECOGNIZER_ENDPOINT")
AZURE_API_KEY = os.getenv("AZURE_FORM_RECOGNIZER_KEY")

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def AzureOCR(document_path,endpoint,key):
    form_recognizer_client = DocumentAnalysisClient(endpoint=endpoint, credential=AzureKeyCredential(key))
    with open(document_path, "rb") as document:
        poller = form_recognizer_client.begin_analyze_document("prebuilt-layout", document)
    result = poller.result()
    return result

def table_exporter(result):
    wb = openpyxl.Workbook()

    if len(result.tables) > 0:
        for table_idx, table in enumerate(result.tables):
            wb.create_sheet(title="page"+"table"+str(table_idx+1))
            sheet = wb["page"+"table"+str(table_idx+1)]
            for cell in table.cells:
                sheet.cell(row=cell.row_index+1,column=cell.column_index+1,value=cell.content)

    emptysheet = wb["Sheet"]
    wb.remove(emptysheet)
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.xlsx')
    wb.save(output_path)
    return output_path

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'
    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        try:
            result = AzureOCR(filepath, AZURE_API_ENDPOINT, AZURE_API_KEY)
        except Exception as e:
            return f"Azure OCR error: {str(e)}"

        excel_path = table_exporter(result)
        
        return redirect(url_for('download_file', filename=os.path.basename(excel_path)))
        
@app.route('/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)
        
if __name__ == '__main__':
    app.run(debug=True)